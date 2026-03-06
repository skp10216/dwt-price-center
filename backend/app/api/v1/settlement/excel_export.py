"""
정산 도메인 — 엑셀 내보내기 (openpyxl)
미수/미지급 리스트를 업체 전달용 전문 양식으로 생성
"""

import io
from typing import Optional, Literal
from datetime import date, datetime
from decimal import Decimal
from urllib.parse import quote

from uuid import UUID

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, text
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

from app.core.database import get_db
from app.api.deps import get_settlement_user
from app.models.user import User
from app.models.voucher import Voucher
from app.models.counterparty import Counterparty
from app.models.receipt import Receipt
from app.models.payment import Payment
from app.models.counterparty_transaction import CounterpartyTransaction
from app.models.enums import (
    VoucherType, TransactionType, TransactionStatus,
    SettlementStatus, PaymentStatus,
)

router = APIRouter()

# ── 스타일 상수 ──
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_HEADER_FILL = PatternFill(start_color="2E3B4E", end_color="2E3B4E", fill_type="solid")
_HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
_TITLE_FONT = Font(name="맑은 고딕", bold=True, size=14)
_SUBTITLE_FONT = Font(name="맑은 고딕", size=10, color="666666")
_DATA_FONT = Font(name="맑은 고딕", size=10)
_SUM_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_SUM_FONT = Font(name="맑은 고딕", bold=True, size=10)
_CURRENCY_FMT = '#,##0'
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _apply_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


async def _query_receivables(db: AsyncSession, date_from=None, date_to=None, search=None):
    """미수 데이터 조회 (페이징 없이 전체)"""
    _active = CounterpartyTransaction.status.notin_([
        TransactionStatus.CANCELLED, TransactionStatus.HIDDEN,
    ])

    v_filters = [Voucher.voucher_type == VoucherType.SALES]
    if date_from:
        v_filters.append(Voucher.trade_date >= date_from)
    if date_to:
        v_filters.append(Voucher.trade_date <= date_to)

    voucher_id_filter = select(Voucher.id).where(*v_filters)

    voucher_sub = (
        select(
            Voucher.counterparty_id,
            func.coalesce(func.sum(Voucher.total_amount), 0).label("total_amount"),
            func.count(Voucher.id).label("voucher_count"),
        ).where(*v_filters).group_by(Voucher.counterparty_id)
    ).subquery()

    receipt_sub = (
        select(
            Voucher.counterparty_id,
            func.coalesce(func.sum(Receipt.amount), 0).label("received"),
        ).join(Voucher, Receipt.voucher_id == Voucher.id)
        .where(Receipt.voucher_id.in_(voucher_id_filter))
        .group_by(Voucher.counterparty_id)
    ).subquery()

    dep_filters = [
        CounterpartyTransaction.transaction_type == TransactionType.DEPOSIT,
        _active,
    ]
    if date_from:
        dep_filters.append(CounterpartyTransaction.transaction_date >= date_from)
    if date_to:
        dep_filters.append(CounterpartyTransaction.transaction_date <= date_to)

    deposit_sub = (
        select(
            CounterpartyTransaction.counterparty_id,
            func.coalesce(func.sum(CounterpartyTransaction.amount), 0).label("txn_received"),
        ).where(*dep_filters).group_by(CounterpartyTransaction.counterparty_id)
    ).subquery()

    total_received = (
        func.coalesce(receipt_sub.c.received, 0) +
        func.coalesce(deposit_sub.c.txn_received, 0)
    )
    balance = func.coalesce(voucher_sub.c.total_amount, 0) - total_received

    query = (
        select(
            Counterparty.name.label("counterparty_name"),
            func.coalesce(voucher_sub.c.total_amount, 0).label("total_amount"),
            func.coalesce(voucher_sub.c.voucher_count, 0).label("voucher_count"),
            total_received.label("total_received"),
            balance.label("balance"),
        )
        .outerjoin(voucher_sub, Counterparty.id == voucher_sub.c.counterparty_id)
        .outerjoin(receipt_sub, Counterparty.id == receipt_sub.c.counterparty_id)
        .outerjoin(deposit_sub, Counterparty.id == deposit_sub.c.counterparty_id)
        .where(
            (voucher_sub.c.counterparty_id.isnot(None)) |
            (deposit_sub.c.counterparty_id.isnot(None))
        )
        .where(balance > 0)
    )
    if search:
        query = query.where(Counterparty.name.ilike(f"%{search}%"))

    query = query.order_by(Counterparty.name)
    result = await db.execute(query)
    return result.all()


async def _query_payables(db: AsyncSession, date_from=None, date_to=None, search=None):
    """미지급 데이터 조회 (페이징 없이 전체)"""
    _active = CounterpartyTransaction.status.notin_([
        TransactionStatus.CANCELLED, TransactionStatus.HIDDEN,
    ])

    v_filters = [Voucher.voucher_type == VoucherType.PURCHASE]
    if date_from:
        v_filters.append(Voucher.trade_date >= date_from)
    if date_to:
        v_filters.append(Voucher.trade_date <= date_to)

    voucher_id_filter = select(Voucher.id).where(*v_filters)

    voucher_sub = (
        select(
            Voucher.counterparty_id,
            func.coalesce(func.sum(Voucher.total_amount), 0).label("total_amount"),
            func.count(Voucher.id).label("voucher_count"),
        ).where(*v_filters).group_by(Voucher.counterparty_id)
    ).subquery()

    payment_sub = (
        select(
            Voucher.counterparty_id,
            func.coalesce(func.sum(Payment.amount), 0).label("paid"),
        ).join(Voucher, Payment.voucher_id == Voucher.id)
        .where(Payment.voucher_id.in_(voucher_id_filter))
        .group_by(Voucher.counterparty_id)
    ).subquery()

    wd_filters = [
        CounterpartyTransaction.transaction_type == TransactionType.WITHDRAWAL,
        _active,
    ]
    if date_from:
        wd_filters.append(CounterpartyTransaction.transaction_date >= date_from)
    if date_to:
        wd_filters.append(CounterpartyTransaction.transaction_date <= date_to)

    withdrawal_sub = (
        select(
            CounterpartyTransaction.counterparty_id,
            func.coalesce(func.sum(CounterpartyTransaction.amount), 0).label("txn_paid"),
        ).where(*wd_filters).group_by(CounterpartyTransaction.counterparty_id)
    ).subquery()

    total_paid = (
        func.coalesce(payment_sub.c.paid, 0) +
        func.coalesce(withdrawal_sub.c.txn_paid, 0)
    )
    balance = func.coalesce(voucher_sub.c.total_amount, 0) - total_paid

    query = (
        select(
            Counterparty.name.label("counterparty_name"),
            func.coalesce(voucher_sub.c.total_amount, 0).label("total_amount"),
            func.coalesce(voucher_sub.c.voucher_count, 0).label("voucher_count"),
            total_paid.label("total_paid"),
            balance.label("balance"),
        )
        .outerjoin(voucher_sub, Counterparty.id == voucher_sub.c.counterparty_id)
        .outerjoin(payment_sub, Counterparty.id == payment_sub.c.counterparty_id)
        .outerjoin(withdrawal_sub, Counterparty.id == withdrawal_sub.c.counterparty_id)
        .where(
            (voucher_sub.c.counterparty_id.isnot(None)) |
            (withdrawal_sub.c.counterparty_id.isnot(None))
        )
        .where(balance > 0)
    )
    if search:
        query = query.where(Counterparty.name.ilike(f"%{search}%"))

    query = query.order_by(Counterparty.name)
    result = await db.execute(query)
    return result.all()


def _build_workbook(
    rows,
    report_type: Literal["receivables", "payables"],
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> Workbook:
    """전문 양식 엑셀 워크북 생성"""
    wb = Workbook()
    ws = wb.active

    is_recv = report_type == "receivables"
    ws.title = "미수 현황" if is_recv else "미지급 현황"

    # ── 열 너비 ──
    col_widths = [6, 30, 10, 20, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # ── 제목 행 ──
    row_num = 1
    _apply_cell(ws, row_num, 1, "미수 현황표" if is_recv else "미지급 현황표",
                font=_TITLE_FONT, alignment=_LEFT)
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
    ws.row_dimensions[row_num].height = 30

    # ── 기간 / 출력일 ──
    row_num = 2
    period_text = "기간: 전체"
    if date_from and date_to:
        period_text = f"기간: {date_from.strftime('%Y-%m-%d')} ~ {date_to.strftime('%Y-%m-%d')}"
    elif date_from:
        period_text = f"기간: {date_from.strftime('%Y-%m-%d')} ~"
    elif date_to:
        period_text = f"기간: ~ {date_to.strftime('%Y-%m-%d')}"
    _apply_cell(ws, row_num, 1, period_text, font=_SUBTITLE_FONT, alignment=_LEFT)
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)

    print_date = f"출력일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    _apply_cell(ws, row_num, 4, print_date, font=_SUBTITLE_FONT, alignment=_RIGHT)
    ws.merge_cells(start_row=row_num, start_column=4, end_row=row_num, end_column=6)
    ws.row_dimensions[row_num].height = 20

    # ── 빈 행 ──
    row_num = 3
    ws.row_dimensions[row_num].height = 8

    # ── 헤더 행 ──
    row_num = 4
    headers = [
        ("No.", _CENTER),
        ("거래처명", _LEFT),
        ("전표 수", _CENTER),
        ("총 매출" if is_recv else "총 매입", _RIGHT),
        ("입금 완료" if is_recv else "출금 완료", _RIGHT),
        ("미수 잔액" if is_recv else "미지급 잔액", _RIGHT),
    ]
    for col, (label, align) in enumerate(headers, 1):
        _apply_cell(ws, row_num, col, label,
                    font=_HEADER_FONT, fill=_HEADER_FILL,
                    alignment=align, border=_THIN_BORDER)
    ws.row_dimensions[row_num].height = 28

    # ── 데이터 행 ──
    sum_vouchers = 0
    sum_total = Decimal("0")
    sum_paid = Decimal("0")
    sum_balance = Decimal("0")

    for idx, r in enumerate(rows, 1):
        row_num += 1
        total_amount = Decimal(str(r.total_amount))
        paid = Decimal(str(r.total_received if is_recv else r.total_paid))
        balance = Decimal(str(r.balance))
        v_count = int(r.voucher_count)

        sum_vouchers += v_count
        sum_total += total_amount
        sum_paid += paid
        sum_balance += balance

        # 짝수/홀수 행 배경색
        row_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") if idx % 2 == 0 else None

        _apply_cell(ws, row_num, 1, idx, font=_DATA_FONT, alignment=_CENTER, border=_THIN_BORDER, fill=row_fill)
        _apply_cell(ws, row_num, 2, r.counterparty_name, font=_DATA_FONT, alignment=_LEFT, border=_THIN_BORDER, fill=row_fill)
        _apply_cell(ws, row_num, 3, v_count, font=_DATA_FONT, alignment=_CENTER, border=_THIN_BORDER, fill=row_fill)
        _apply_cell(ws, row_num, 4, float(total_amount), font=_DATA_FONT, alignment=_RIGHT,
                    border=_THIN_BORDER, number_format=_CURRENCY_FMT, fill=row_fill)
        _apply_cell(ws, row_num, 5, float(paid), font=_DATA_FONT, alignment=_RIGHT,
                    border=_THIN_BORDER, number_format=_CURRENCY_FMT, fill=row_fill)
        _apply_cell(ws, row_num, 6, float(balance), font=Font(name="맑은 고딕", bold=True, size=10, color="C0392B" if is_recv else "2E86C1"),
                    alignment=_RIGHT, border=_THIN_BORDER, number_format=_CURRENCY_FMT, fill=row_fill)

    # ── 합계 행 ──
    row_num += 1
    _apply_cell(ws, row_num, 1, "", font=_SUM_FONT, fill=_SUM_FILL, border=_THIN_BORDER, alignment=_CENTER)
    _apply_cell(ws, row_num, 2, f"합계 ({len(rows)}개 거래처)", font=_SUM_FONT, fill=_SUM_FILL, border=_THIN_BORDER, alignment=_LEFT)
    _apply_cell(ws, row_num, 3, sum_vouchers, font=_SUM_FONT, fill=_SUM_FILL, border=_THIN_BORDER, alignment=_CENTER)
    _apply_cell(ws, row_num, 4, float(sum_total), font=_SUM_FONT, fill=_SUM_FILL,
                border=_THIN_BORDER, alignment=_RIGHT, number_format=_CURRENCY_FMT)
    _apply_cell(ws, row_num, 5, float(sum_paid), font=_SUM_FONT, fill=_SUM_FILL,
                border=_THIN_BORDER, alignment=_RIGHT, number_format=_CURRENCY_FMT)
    _apply_cell(ws, row_num, 6, float(sum_balance),
                font=Font(name="맑은 고딕", bold=True, size=11, color="C0392B" if is_recv else "2E86C1"),
                fill=_SUM_FILL, border=_THIN_BORDER, alignment=_RIGHT, number_format=_CURRENCY_FMT)
    ws.row_dimensions[row_num].height = 28

    # ── 인쇄 설정 ──
    ws.print_title_rows = "4:4"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"

    return wb


@router.get("/receivables/export")
async def export_receivables_excel(
    search: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_settlement_user),
):
    """미수 현황 엑셀 다운로드 (업체 전달용 전문 양식)"""
    rows = await _query_receivables(db, date_from, date_to, search)
    wb = _build_workbook(rows, "receivables", date_from, date_to)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = datetime.now().strftime("%Y%m%d")
    filename = f"receivables_{today}.xlsx"
    filename_display = quote(f"미수현황_{today}.xlsx")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}; filename*=UTF-8''{filename_display}"},
    )


@router.get("/payables/export")
async def export_payables_excel(
    search: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_settlement_user),
):
    """미지급 현황 엑셀 다운로드 (업체 전달용 전문 양식)"""
    rows = await _query_payables(db, date_from, date_to, search)
    wb = _build_workbook(rows, "payables", date_from, date_to)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = datetime.now().strftime("%Y%m%d")
    filename = f"payables_{today}.xlsx"
    filename_display = quote(f"미지급현황_{today}.xlsx")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}; filename*=UTF-8''{filename_display}"},
    )


# ============================================================================
# 거래처 상세 미수/미지급 엑셀 (전표별 내역)
# ============================================================================

_SETTLEMENT_STATUS_KO = {
    "OPEN": "미정산", "SETTLING": "정산중", "SETTLED": "정산완료", "LOCKED": "마감",
}
_PAYMENT_STATUS_KO = {
    "UNPAID": "미지급", "PARTIAL": "부분지급", "PAID": "지급완료", "LOCKED": "마감",
}


def _build_counterparty_workbook(
    counterparty_name: str,
    vouchers: list,
    report_type: Literal["receivables", "payables"],
    summary_data: dict,
) -> Workbook:
    """거래처 상세 전표별 미수/미지급 엑셀"""
    wb = Workbook()
    ws = wb.active
    is_recv = report_type == "receivables"
    ws.title = f"{'미수' if is_recv else '미지급'} 상세"

    # 열 너비
    col_widths = [6, 12, 18, 8, 15, 15, 15, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # ── 제목 ──
    row_num = 1
    title = f"{counterparty_name} — {'미수' if is_recv else '미지급'} 상세 내역"
    _apply_cell(ws, row_num, 1, title, font=_TITLE_FONT, alignment=_LEFT)
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
    ws.row_dimensions[row_num].height = 30

    # ── 요약 정보 ──
    row_num = 2
    total_label = "총 매출" if is_recv else "총 매입"
    paid_label = "입금 완료" if is_recv else "출금 완료"
    balance_label = "미수 잔액" if is_recv else "미지급 잔액"
    total_amt = summary_data.get("total_amount", 0)
    paid_amt = summary_data.get("paid_amount", 0)
    balance_amt = summary_data.get("balance", 0)

    summary_text = f"{total_label}: {total_amt:,.0f}원  |  {paid_label}: {paid_amt:,.0f}원  |  {balance_label}: {balance_amt:,.0f}원"
    _apply_cell(ws, row_num, 1, summary_text, font=_SUBTITLE_FONT, alignment=_LEFT)
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)

    print_date = f"출력일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    _apply_cell(ws, row_num, 8, print_date, font=_SUBTITLE_FONT, alignment=_RIGHT)
    ws.merge_cells(start_row=row_num, start_column=8, end_row=row_num, end_column=9)
    ws.row_dimensions[row_num].height = 20

    # ── 빈 행 ──
    row_num = 3
    ws.row_dimensions[row_num].height = 8

    # ── 헤더 ──
    row_num = 4
    headers = [
        ("No.", _CENTER),
        ("거래일", _CENTER),
        ("전표번호", _LEFT),
        ("수량", _CENTER),
        ("전표 금액", _RIGHT),
        ("수금액" if is_recv else "지급액", _RIGHT),
        ("잔액", _RIGHT),
        ("정산상태", _CENTER),
        ("지급상태", _CENTER),
    ]
    for col, (label, align) in enumerate(headers, 1):
        _apply_cell(ws, row_num, col, label,
                    font=_HEADER_FONT, fill=_HEADER_FILL,
                    alignment=align, border=_THIN_BORDER)
    ws.row_dimensions[row_num].height = 28

    # ── 데이터 ──
    sum_qty = 0
    sum_amount = Decimal("0")
    sum_paid = Decimal("0")
    sum_balance = Decimal("0")

    for idx, v in enumerate(vouchers, 1):
        row_num += 1
        amount = Decimal(str(v["total_amount"]))
        paid = Decimal(str(v["paid"]))
        bal = Decimal(str(v["balance"]))
        qty = int(v["quantity"])

        sum_qty += qty
        sum_amount += amount
        sum_paid += paid
        sum_balance += bal

        row_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") if idx % 2 == 0 else None

        _apply_cell(ws, row_num, 1, idx, font=_DATA_FONT, alignment=_CENTER, border=_THIN_BORDER, fill=row_fill)
        _apply_cell(ws, row_num, 2, str(v["trade_date"]), font=_DATA_FONT, alignment=_CENTER, border=_THIN_BORDER, fill=row_fill)
        _apply_cell(ws, row_num, 3, v["voucher_number"], font=_DATA_FONT, alignment=_LEFT, border=_THIN_BORDER, fill=row_fill)
        _apply_cell(ws, row_num, 4, qty, font=_DATA_FONT, alignment=_CENTER, border=_THIN_BORDER, fill=row_fill)
        _apply_cell(ws, row_num, 5, float(amount), font=_DATA_FONT, alignment=_RIGHT,
                    border=_THIN_BORDER, number_format=_CURRENCY_FMT, fill=row_fill)
        _apply_cell(ws, row_num, 6, float(paid), font=_DATA_FONT, alignment=_RIGHT,
                    border=_THIN_BORDER, number_format=_CURRENCY_FMT, fill=row_fill)

        bal_color = "C0392B" if is_recv else "2E86C1"
        _apply_cell(ws, row_num, 7, float(bal),
                    font=Font(name="맑은 고딕", bold=True, size=10, color=bal_color if bal > 0 else "999999"),
                    alignment=_RIGHT, border=_THIN_BORDER, number_format=_CURRENCY_FMT, fill=row_fill)

        s_key = v["settlement_status"].upper() if isinstance(v["settlement_status"], str) else v["settlement_status"]
        p_key = v["payment_status"].upper() if isinstance(v["payment_status"], str) else v["payment_status"]
        s_status = _SETTLEMENT_STATUS_KO.get(s_key, v["settlement_status"])
        p_status = _PAYMENT_STATUS_KO.get(p_key, v["payment_status"])
        _apply_cell(ws, row_num, 8, s_status, font=_DATA_FONT, alignment=_CENTER, border=_THIN_BORDER, fill=row_fill)
        _apply_cell(ws, row_num, 9, p_status, font=_DATA_FONT, alignment=_CENTER, border=_THIN_BORDER, fill=row_fill)

    # ── 합계 ──
    row_num += 1
    _apply_cell(ws, row_num, 1, "", font=_SUM_FONT, fill=_SUM_FILL, border=_THIN_BORDER, alignment=_CENTER)
    _apply_cell(ws, row_num, 2, f"합계 ({len(vouchers)}건)", font=_SUM_FONT, fill=_SUM_FILL, border=_THIN_BORDER, alignment=_LEFT)
    _apply_cell(ws, row_num, 3, "", font=_SUM_FONT, fill=_SUM_FILL, border=_THIN_BORDER, alignment=_LEFT)
    _apply_cell(ws, row_num, 4, sum_qty, font=_SUM_FONT, fill=_SUM_FILL, border=_THIN_BORDER, alignment=_CENTER)
    _apply_cell(ws, row_num, 5, float(sum_amount), font=_SUM_FONT, fill=_SUM_FILL,
                border=_THIN_BORDER, alignment=_RIGHT, number_format=_CURRENCY_FMT)
    _apply_cell(ws, row_num, 6, float(sum_paid), font=_SUM_FONT, fill=_SUM_FILL,
                border=_THIN_BORDER, alignment=_RIGHT, number_format=_CURRENCY_FMT)
    bal_color = "C0392B" if is_recv else "2E86C1"
    _apply_cell(ws, row_num, 7, float(sum_balance),
                font=Font(name="맑은 고딕", bold=True, size=11, color=bal_color),
                fill=_SUM_FILL, border=_THIN_BORDER, alignment=_RIGHT, number_format=_CURRENCY_FMT)
    _apply_cell(ws, row_num, 8, "", font=_SUM_FONT, fill=_SUM_FILL, border=_THIN_BORDER, alignment=_CENTER)
    _apply_cell(ws, row_num, 9, "", font=_SUM_FONT, fill=_SUM_FILL, border=_THIN_BORDER, alignment=_CENTER)
    ws.row_dimensions[row_num].height = 28

    # ── 인쇄 설정 ──
    ws.print_title_rows = "4:4"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"

    return wb


@router.get("/counterparty/{counterparty_id}/export")
async def export_counterparty_detail_excel(
    counterparty_id: UUID,
    voucher_type: str = Query("sales", description="sales 또는 purchase"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_settlement_user),
):
    """거래처 상세 미수/미지급 엑셀 — 전표별 내역 (업체 전달용)"""
    cp = (await db.execute(
        select(Counterparty.name).where(Counterparty.id == counterparty_id)
    )).scalar_one_or_none()
    if not cp:
        raise HTTPException(status_code=404, detail="거래처를 찾을 수 없습니다")
    cp_name = cp

    is_sales = voucher_type == "sales"
    v_type = VoucherType.SALES if is_sales else VoucherType.PURCHASE

    # 레거시 입금/송금 서브쿼리
    if is_sales:
        legacy_sub = (
            select(
                Receipt.voucher_id,
                func.coalesce(func.sum(Receipt.amount), 0).label("legacy_paid"),
            ).group_by(Receipt.voucher_id)
        ).subquery()
    else:
        legacy_sub = (
            select(
                Payment.voucher_id,
                func.coalesce(func.sum(Payment.amount), 0).label("legacy_paid"),
            ).group_by(Payment.voucher_id)
        ).subquery()

    # 배분 합계 서브쿼리
    from app.models.transaction_allocation import TransactionAllocation
    alloc_sub = (
        select(
            TransactionAllocation.voucher_id,
            func.coalesce(func.sum(TransactionAllocation.allocated_amount), 0).label("alloc_paid"),
        ).group_by(TransactionAllocation.voucher_id)
    ).subquery()

    paid_expr = (
        func.coalesce(legacy_sub.c.legacy_paid, 0) +
        func.coalesce(alloc_sub.c.alloc_paid, 0)
    )

    query = (
        select(
            Voucher.trade_date,
            Voucher.voucher_number,
            Voucher.quantity,
            Voucher.total_amount,
            paid_expr.label("paid"),
            (Voucher.total_amount - paid_expr).label("balance"),
            Voucher.settlement_status,
            Voucher.payment_status,
        )
        .outerjoin(legacy_sub, Voucher.id == legacy_sub.c.voucher_id)
        .outerjoin(alloc_sub, Voucher.id == alloc_sub.c.voucher_id)
        .where(
            Voucher.counterparty_id == counterparty_id,
            Voucher.voucher_type == v_type,
        )
        .order_by(Voucher.trade_date.desc(), Voucher.voucher_number)
    )
    result = await db.execute(query)
    rows = result.all()

    voucher_list = []
    total_amount = Decimal("0")
    total_paid = Decimal("0")
    total_balance = Decimal("0")
    for r in rows:
        amt = Decimal(str(r.total_amount))
        paid = Decimal(str(r.paid))
        bal = Decimal(str(r.balance))
        total_amount += amt
        total_paid += paid
        total_balance += bal
        voucher_list.append({
            "trade_date": r.trade_date,
            "voucher_number": r.voucher_number,
            "quantity": r.quantity,
            "total_amount": amt,
            "paid": paid,
            "balance": bal,
            "settlement_status": r.settlement_status.value if hasattr(r.settlement_status, 'value') else str(r.settlement_status),
            "payment_status": r.payment_status.value if hasattr(r.payment_status, 'value') else str(r.payment_status),
        })

    report_type = "receivables" if is_sales else "payables"
    summary = {
        "total_amount": float(total_amount),
        "paid_amount": float(total_paid),
        "balance": float(total_balance),
    }

    wb = _build_counterparty_workbook(cp_name, voucher_list, report_type, summary)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = datetime.now().strftime("%Y%m%d")
    type_tag = "recv" if is_sales else "pay"
    filename = f"{type_tag}_{today}.xlsx"
    display_label = "미수" if is_sales else "미지급"
    filename_display = quote(f"{display_label}상세_{cp_name}_{today}.xlsx")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}; filename*=UTF-8''{filename_display}"},
    )
