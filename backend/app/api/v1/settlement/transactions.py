"""
정산 도메인 - 거래처 입출금 이벤트 관리
전표 무관하게 거래처 수준에서 입출금을 기록하고, 전표에 자동/수동 배분
"""

from uuid import UUID
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import List, Optional
import io
import re

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile, status
from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.api.deps import get_current_user
from app.models.user import User
from app.models.counterparty import Counterparty, CounterpartyAlias
from app.models.voucher import Voucher
from app.models.counterparty_transaction import CounterpartyTransaction
from app.models.corporate_entity import CorporateEntity
from app.models.bank_import import BankImportJob, BankImportLine
from app.models.transaction_allocation import TransactionAllocation
from app.models.receipt import Receipt
from app.models.payment import Payment
from app.models.audit_log import AuditLog
from app.models.enums import (
    TransactionType, TransactionSource, TransactionStatus,
    VoucherType, SettlementStatus, PaymentStatus, AuditAction,
)
from app.schemas.settlement import (
    TransactionCreate, TransactionUpdate, TransactionResponse,
    TransactionDetailResponse, TransactionListResponse,
    AllocationRequest, AutoAllocateRequest, AllocationResponse,
    CounterpartyTimelineItem, CounterpartyBalanceSummary,
    TransactionHoldRequest, TransactionHideRequest,
)

router = APIRouter()


# =============================================================================
# 헬퍼 함수
# =============================================================================

async def _get_voucher_allocated_amount(voucher_id: UUID, db: AsyncSession) -> Decimal:
    """전표에 이미 배분된 총액 (TransactionAllocation 기준)"""
    result = await db.execute(
        select(func.coalesce(func.sum(TransactionAllocation.allocated_amount), 0))
        .where(TransactionAllocation.voucher_id == voucher_id)
    )
    return result.scalar() or Decimal("0")


async def _get_voucher_legacy_amount(voucher_id: UUID, voucher_type: str, db: AsyncSession) -> Decimal:
    """레거시 Receipt/Payment 합계 (전환기 호환)"""
    if voucher_type == VoucherType.SALES.value or voucher_type == VoucherType.SALES:
        result = await db.execute(
            select(func.coalesce(func.sum(Receipt.amount), 0))
            .where(Receipt.voucher_id == voucher_id)
        )
    else:
        result = await db.execute(
            select(func.coalesce(func.sum(Payment.amount), 0))
            .where(Payment.voucher_id == voucher_id)
        )
    return result.scalar() or Decimal("0")


async def _update_voucher_status(voucher_id: UUID, db: AsyncSession) -> None:
    """배분 총액 기반으로 전표 상태 자동 전이"""
    v = await db.get(Voucher, voucher_id)
    if not v or v.settlement_status == SettlementStatus.LOCKED or v.payment_status == PaymentStatus.LOCKED:
        return

    alloc_total = await _get_voucher_allocated_amount(voucher_id, db)
    legacy_total = await _get_voucher_legacy_amount(voucher_id, v.voucher_type, db)
    total_settled = alloc_total + legacy_total

    if v.voucher_type == VoucherType.SALES:
        if total_settled >= v.total_amount:
            v.settlement_status = SettlementStatus.SETTLED
        elif total_settled > 0:
            v.settlement_status = SettlementStatus.SETTLING
        else:
            v.settlement_status = SettlementStatus.OPEN
    else:  # PURCHASE
        if total_settled >= v.total_amount:
            v.payment_status = PaymentStatus.PAID
        elif total_settled > 0:
            v.payment_status = PaymentStatus.PARTIAL
        else:
            v.payment_status = PaymentStatus.UNPAID


async def _update_transaction_status(txn: CounterpartyTransaction, db: AsyncSession) -> None:
    """배분 누적에 따라 Transaction 상태 전이"""
    alloc_sum = (await db.execute(
        select(func.coalesce(func.sum(TransactionAllocation.allocated_amount), 0))
        .where(TransactionAllocation.transaction_id == txn.id)
    )).scalar() or Decimal("0")

    txn.allocated_amount = alloc_sum

    if alloc_sum >= txn.amount:
        txn.status = TransactionStatus.ALLOCATED
    elif alloc_sum > 0:
        txn.status = TransactionStatus.PARTIAL
    else:
        txn.status = TransactionStatus.PENDING


def _txn_to_response(
    txn: CounterpartyTransaction,
    counterparty_name: str = None,
    corporate_entity_name: str = None,
    bank_name: str = None,
    account_number: str = None,
) -> TransactionResponse:
    """Transaction → Response 변환"""
    return TransactionResponse(
        id=txn.id,
        counterparty_id=txn.counterparty_id,
        counterparty_name=counterparty_name,
        transaction_type=txn.transaction_type.value if hasattr(txn.transaction_type, 'value') else txn.transaction_type,
        transaction_date=txn.transaction_date,
        amount=txn.amount,
        allocated_amount=txn.allocated_amount,
        unallocated_amount=txn.amount - txn.allocated_amount,
        memo=txn.memo,
        source=txn.source.value if hasattr(txn.source, 'value') else txn.source,
        bank_reference=txn.bank_reference,
        netting_record_id=txn.netting_record_id,
        corporate_entity_id=txn.corporate_entity_id,
        corporate_entity_name=corporate_entity_name,
        bank_name=bank_name,
        account_number=account_number,
        status=txn.status.value if hasattr(txn.status, 'value') else txn.status,
        created_by=txn.created_by,
        created_at=txn.created_at,
        updated_at=txn.updated_at,
    )


# =============================================================================
# CRUD 엔드포인트
# =============================================================================

@router.get("", response_model=TransactionListResponse)
async def list_transactions(
    counterparty_id: Optional[UUID] = None,
    transaction_type: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    source: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    search: Optional[str] = Query(None, description="거래처명 + 메모 통합 검색"),
    corporate_entity_id: Optional[UUID] = Query(None, description="법인 필터"),
    amount_min: Optional[Decimal] = Query(None, description="최소 금액"),
    amount_max: Optional[Decimal] = Query(None, description="최대 금액"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """입출금 이벤트 목록 조회 (검색/금액범위/복수상태 지원)"""
    query = select(CounterpartyTransaction).join(Counterparty)
    count_query = select(func.count(CounterpartyTransaction.id)).join(Counterparty)

    filters = []
    if counterparty_id:
        filters.append(CounterpartyTransaction.counterparty_id == counterparty_id)
    if transaction_type:
        filters.append(CounterpartyTransaction.transaction_type == transaction_type)

    # 복수 상태 지원 (쉼표 구분)
    if status_filter:
        statuses = [s.strip() for s in status_filter.split(",") if s.strip()]
        if len(statuses) == 1:
            filters.append(CounterpartyTransaction.status == statuses[0])
        else:
            filters.append(CounterpartyTransaction.status.in_(statuses))

    if source:
        filters.append(CounterpartyTransaction.source == source)
    if corporate_entity_id:
        filters.append(CounterpartyTransaction.corporate_entity_id == corporate_entity_id)
    if date_from:
        filters.append(CounterpartyTransaction.transaction_date >= date_from)
    if date_to:
        filters.append(CounterpartyTransaction.transaction_date <= date_to)

    # 금액 범위 필터
    if amount_min is not None:
        filters.append(CounterpartyTransaction.amount >= amount_min)
    if amount_max is not None:
        filters.append(CounterpartyTransaction.amount <= amount_max)

    # 거래처명 + 메모 통합 검색
    if search:
        search_term = f"%{search}%"
        filters.append(
            (Counterparty.name.ilike(search_term)) |
            (CounterpartyTransaction.memo.ilike(search_term))
        )

    # 기본 제외: cancelled, hidden (명시적 필터 시에만 포함)
    if not status_filter:
        filters.append(CounterpartyTransaction.status.notin_([
            TransactionStatus.CANCELLED,
            TransactionStatus.HIDDEN,
        ]))

    if filters:
        query = query.where(and_(*filters))
        count_query = count_query.where(and_(*filters))

    total = (await db.execute(count_query)).scalar() or 0

    query = query.order_by(
        CounterpartyTransaction.transaction_date.desc(),
        CounterpartyTransaction.created_at.desc(),
    )
    query = query.offset((page - 1) * page_size).limit(page_size)

    result = await db.execute(query)
    txns = result.scalars().all()

    # 거래처명 일괄 로드
    cp_ids = {t.counterparty_id for t in txns}
    cp_map = {}
    if cp_ids:
        cp_result = await db.execute(
            select(Counterparty.id, Counterparty.name)
            .where(Counterparty.id.in_(cp_ids))
        )
        cp_map = {row.id: row.name for row in cp_result.all()}

    # 법인명 일괄 로드
    ce_ids = {t.corporate_entity_id for t in txns if t.corporate_entity_id}
    ce_map = {}
    if ce_ids:
        ce_result = await db.execute(
            select(CorporateEntity.id, CorporateEntity.name)
            .where(CorporateEntity.id.in_(ce_ids))
        )
        ce_map = {row.id: row.name for row in ce_result.all()}

    # 은행 정보 일괄 로드 (bank_import_line_id → BankImportJob)
    bil_ids = {t.bank_import_line_id for t in txns if t.bank_import_line_id}
    bank_map: dict = {}  # {bank_import_line_id: (bank_name, account_number)}
    if bil_ids:
        bank_result = await db.execute(
            select(BankImportLine.id, BankImportJob.bank_name, BankImportJob.account_number)
            .join(BankImportJob, BankImportLine.import_job_id == BankImportJob.id)
            .where(BankImportLine.id.in_(bil_ids))
        )
        bank_map = {row.id: (row.bank_name, row.account_number) for row in bank_result.all()}

    def _build_response(t):
        bank_info = bank_map.get(t.bank_import_line_id, (None, None))
        return _txn_to_response(
            t,
            counterparty_name=cp_map.get(t.counterparty_id),
            corporate_entity_name=ce_map.get(t.corporate_entity_id),
            bank_name=bank_info[0],
            account_number=bank_info[1],
        )

    return TransactionListResponse(
        transactions=[_build_response(t) for t in txns],
        total=total,
        page=page,
        page_size=page_size,
    )


@router.post("", response_model=TransactionResponse, status_code=201)
async def create_transaction(
    data: TransactionCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """수동 입출금 이벤트 등록"""
    cp = await db.get(Counterparty, data.counterparty_id)
    if not cp:
        raise HTTPException(status_code=404, detail="거래처를 찾을 수 없습니다")
    if not cp.is_active:
        raise HTTPException(status_code=400, detail="비활성 거래처에는 입출금을 등록할 수 없습니다")

    txn = CounterpartyTransaction(
        counterparty_id=data.counterparty_id,
        transaction_type=data.transaction_type,
        transaction_date=data.transaction_date,
        amount=data.amount,
        memo=data.memo,
        source=TransactionSource.MANUAL,
        bank_reference=data.bank_reference,
        status=TransactionStatus.PENDING,
        created_by=current_user.id,
    )
    db.add(txn)
    await db.flush()

    db.add(AuditLog(
        user_id=current_user.id,
        action=AuditAction.TRANSACTION_CREATE,
        target_type="counterparty_transaction",
        target_id=txn.id,
        after_data={
            "counterparty_id": str(data.counterparty_id),
            "type": data.transaction_type,
            "amount": str(data.amount),
            "date": str(data.transaction_date),
        },
    ))

    return _txn_to_response(txn, cp.name)


@router.get("/{transaction_id}", response_model=TransactionDetailResponse)
async def get_transaction(
    transaction_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """입출금 이벤트 상세 (배분 내역 포함)"""
    result = await db.execute(
        select(CounterpartyTransaction)
        .options(selectinload(CounterpartyTransaction.allocations))
        .where(CounterpartyTransaction.id == transaction_id)
    )
    txn = result.scalar_one_or_none()
    if not txn:
        raise HTTPException(status_code=404, detail="입출금 이벤트를 찾을 수 없습니다")

    cp = await db.get(Counterparty, txn.counterparty_id)

    # 법인명 로드
    ce_name = None
    if txn.corporate_entity_id:
        ce = await db.get(CorporateEntity, txn.corporate_entity_id)
        ce_name = ce.name if ce else None

    # 은행 정보 로드 (bank_import_line → job)
    b_name, b_account = None, None
    if txn.bank_import_line_id:
        bil_result = await db.execute(
            select(BankImportJob.bank_name, BankImportJob.account_number)
            .join(BankImportLine, BankImportLine.import_job_id == BankImportJob.id)
            .where(BankImportLine.id == txn.bank_import_line_id)
        )
        bil_row = bil_result.first()
        if bil_row:
            b_name, b_account = bil_row.bank_name, bil_row.account_number

    # 배분 내역에 전표 정보 배치 로드
    alloc_voucher_ids = [alloc.voucher_id for alloc in txn.allocations]
    voucher_map = {}
    if alloc_voucher_ids:
        v_result = await db.execute(
            select(Voucher).where(Voucher.id.in_(alloc_voucher_ids))
        )
        voucher_map = {v.id: v for v in v_result.scalars().all()}

    alloc_responses = []
    for alloc in txn.allocations:
        v = voucher_map.get(alloc.voucher_id)
        alloc_responses.append(AllocationResponse(
            id=alloc.id,
            transaction_id=alloc.transaction_id,
            voucher_id=alloc.voucher_id,
            voucher_number=v.voucher_number if v else None,
            voucher_trade_date=v.trade_date if v else None,
            voucher_total_amount=v.total_amount if v else None,
            allocated_amount=alloc.allocated_amount,
            allocation_order=alloc.allocation_order,
            memo=alloc.memo,
            created_at=alloc.created_at,
        ))

    resp = _txn_to_response(txn, cp.name if cp else None, ce_name, b_name, b_account)
    return TransactionDetailResponse(
        **resp.model_dump(),
        allocations=alloc_responses,
    )


@router.patch("/{transaction_id}", response_model=TransactionResponse)
async def update_transaction(
    transaction_id: UUID,
    data: TransactionUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """입출금 이벤트 수정 (PENDING 상태만)"""
    txn = await db.get(CounterpartyTransaction, transaction_id)
    if not txn:
        raise HTTPException(status_code=404, detail="입출금 이벤트를 찾을 수 없습니다")
    if txn.status != TransactionStatus.PENDING:
        raise HTTPException(status_code=400, detail="미배분 상태에서만 수정할 수 있습니다")

    before = {"amount": str(txn.amount), "date": str(txn.transaction_date)}

    if data.transaction_date is not None:
        txn.transaction_date = data.transaction_date
    if data.amount is not None:
        txn.amount = data.amount
    if data.memo is not None:
        txn.memo = data.memo

    db.add(AuditLog(
        user_id=current_user.id,
        action=AuditAction.TRANSACTION_UPDATE,
        target_type="counterparty_transaction",
        target_id=txn.id,
        before_data=before,
        after_data={"amount": str(txn.amount), "date": str(txn.transaction_date)},
    ))

    cp = await db.get(Counterparty, txn.counterparty_id)
    return _txn_to_response(txn, cp.name if cp else None)


@router.delete("/{transaction_id}", status_code=204)
async def cancel_transaction(
    transaction_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """입출금 이벤트 취소 (배분 해제 후 CANCELLED)"""
    txn = await db.get(CounterpartyTransaction, transaction_id)
    if not txn:
        raise HTTPException(status_code=404, detail="입출금 이벤트를 찾을 수 없습니다")
    if txn.status == TransactionStatus.CANCELLED:
        raise HTTPException(status_code=400, detail="이미 취소된 이벤트입니다")

    # 배분 해제
    alloc_result = await db.execute(
        select(TransactionAllocation)
        .where(TransactionAllocation.transaction_id == txn.id)
    )
    allocs = alloc_result.scalars().all()
    affected_voucher_ids = [a.voucher_id for a in allocs]

    for alloc in allocs:
        await db.delete(alloc)

    txn.status = TransactionStatus.CANCELLED
    txn.allocated_amount = Decimal("0")
    await db.flush()

    # 영향받은 전표 상태 재계산
    for vid in affected_voucher_ids:
        await _update_voucher_status(vid, db)

    db.add(AuditLog(
        user_id=current_user.id,
        action=AuditAction.TRANSACTION_CANCEL,
        target_type="counterparty_transaction",
        target_id=txn.id,
        before_data={"status": "allocated", "allocated_amount": str(txn.amount)},
    ))


@router.post("/{transaction_id}/hold", response_model=TransactionResponse)
async def hold_transaction(
    transaction_id: UUID,
    data: TransactionHoldRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """입출금 이벤트 보류 처리 (사유 필수)"""
    txn = await db.get(CounterpartyTransaction, transaction_id)
    if not txn:
        raise HTTPException(status_code=404, detail="입출금 이벤트를 찾을 수 없습니다")
    if txn.status in (TransactionStatus.CANCELLED, TransactionStatus.HIDDEN):
        raise HTTPException(status_code=400, detail="취소/숨김 상태에서는 보류할 수 없습니다")
    if txn.status == TransactionStatus.ON_HOLD:
        raise HTTPException(status_code=400, detail="이미 보류 상태입니다")

    prev_status = txn.status.value if hasattr(txn.status, 'value') else txn.status
    txn.status = TransactionStatus.ON_HOLD

    db.add(AuditLog(
        user_id=current_user.id,
        action=AuditAction.TRANSACTION_HOLD,
        target_type="counterparty_transaction",
        target_id=txn.id,
        before_data={"status": prev_status},
        after_data={"status": "on_hold", "reason": data.reason},
        description=data.reason,
    ))

    cp = await db.get(Counterparty, txn.counterparty_id)
    return _txn_to_response(txn, cp.name if cp else None)


@router.post("/{transaction_id}/unhold", response_model=TransactionResponse)
async def unhold_transaction(
    transaction_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """입출금 이벤트 보류 해제 (배분 상태에 따라 자동 전이)"""
    txn = await db.get(CounterpartyTransaction, transaction_id)
    if not txn:
        raise HTTPException(status_code=404, detail="입출금 이벤트를 찾을 수 없습니다")
    if txn.status != TransactionStatus.ON_HOLD:
        raise HTTPException(status_code=400, detail="보류 상태가 아닙니다")

    # 배분 상태에 따라 자동 전이
    await _update_transaction_status(txn, db)

    db.add(AuditLog(
        user_id=current_user.id,
        action=AuditAction.TRANSACTION_UNHOLD,
        target_type="counterparty_transaction",
        target_id=txn.id,
        before_data={"status": "on_hold"},
        after_data={"status": txn.status.value if hasattr(txn.status, 'value') else txn.status},
    ))

    cp = await db.get(Counterparty, txn.counterparty_id)
    return _txn_to_response(txn, cp.name if cp else None)


@router.post("/{transaction_id}/hide", response_model=TransactionResponse)
async def hide_transaction(
    transaction_id: UUID,
    data: TransactionHideRequest = TransactionHideRequest(),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """입출금 이벤트 숨김 처리 (삭제 대체)"""
    txn = await db.get(CounterpartyTransaction, transaction_id)
    if not txn:
        raise HTTPException(status_code=404, detail="입출금 이벤트를 찾을 수 없습니다")
    if txn.status == TransactionStatus.CANCELLED:
        raise HTTPException(status_code=400, detail="취소된 이벤트는 숨길 수 없습니다")
    if txn.status == TransactionStatus.HIDDEN:
        raise HTTPException(status_code=400, detail="이미 숨김 상태입니다")

    prev_status = txn.status.value if hasattr(txn.status, 'value') else txn.status
    txn.status = TransactionStatus.HIDDEN

    db.add(AuditLog(
        user_id=current_user.id,
        action=AuditAction.TRANSACTION_HIDE,
        target_type="counterparty_transaction",
        target_id=txn.id,
        before_data={"status": prev_status},
        after_data={"status": "hidden", "reason": data.reason},
        description=data.reason,
    ))

    cp = await db.get(Counterparty, txn.counterparty_id)
    return _txn_to_response(txn, cp.name if cp else None)


@router.post("/{transaction_id}/unhide", response_model=TransactionResponse)
async def unhide_transaction(
    transaction_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """입출금 이벤트 숨김 해제 (배분 상태에 따라 자동 전이)"""
    txn = await db.get(CounterpartyTransaction, transaction_id)
    if not txn:
        raise HTTPException(status_code=404, detail="입출금 이벤트를 찾을 수 없습니다")
    if txn.status != TransactionStatus.HIDDEN:
        raise HTTPException(status_code=400, detail="숨김 상태가 아닙니다")

    # 배분 상태에 따라 자동 전이
    await _update_transaction_status(txn, db)

    db.add(AuditLog(
        user_id=current_user.id,
        action=AuditAction.TRANSACTION_UNHIDE,
        target_type="counterparty_transaction",
        target_id=txn.id,
        before_data={"status": "hidden"},
        after_data={"status": txn.status.value if hasattr(txn.status, 'value') else txn.status},
    ))

    cp = await db.get(Counterparty, txn.counterparty_id)
    return _txn_to_response(txn, cp.name if cp else None)


@router.post("/batch-cancel", status_code=200)
async def batch_cancel_transactions(
    transaction_ids: List[UUID] = Body(..., description="취소할 입출금 ID 목록"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """입출금 일괄 취소 (이미 취소된 건은 건너뜀)"""
    cancelled_count = 0
    skipped_count = 0
    errors = []

    for tid in transaction_ids:
        txn = await db.get(CounterpartyTransaction, tid)
        if not txn:
            skipped_count += 1
            errors.append(f"입출금 ID {tid}를 찾을 수 없습니다.")
            continue

        if txn.status == TransactionStatus.CANCELLED:
            skipped_count += 1
            continue

        # 배분 해제
        alloc_result = await db.execute(
            select(TransactionAllocation)
            .where(TransactionAllocation.transaction_id == txn.id)
        )
        allocs = alloc_result.scalars().all()
        affected_voucher_ids = [a.voucher_id for a in allocs]

        for alloc in allocs:
            await db.delete(alloc)

        prev_status = txn.status.value if hasattr(txn.status, 'value') else txn.status
        txn.status = TransactionStatus.CANCELLED
        txn.allocated_amount = Decimal("0")
        await db.flush()

        # 영향받은 전표 상태 재계산
        for vid in affected_voucher_ids:
            await _update_voucher_status(vid, db)

        db.add(AuditLog(
            user_id=current_user.id,
            action=AuditAction.TRANSACTION_CANCEL,
            target_type="counterparty_transaction",
            target_id=txn.id,
            before_data={"status": prev_status, "allocated_amount": str(txn.amount)},
        ))
        cancelled_count += 1

    await db.flush()
    return {
        "cancelled_count": cancelled_count,
        "skipped_count": skipped_count,
        "errors": errors,
    }


# =============================================================================
# 배분 엔드포인트
# =============================================================================

@router.post("/{transaction_id}/auto-allocate", response_model=list[AllocationResponse])
async def auto_allocate(
    transaction_id: UUID,
    data: AutoAllocateRequest = AutoAllocateRequest(),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """FIFO 자동 배분"""
    txn = await db.get(CounterpartyTransaction, transaction_id)
    if not txn:
        raise HTTPException(status_code=404, detail="입출금 이벤트를 찾을 수 없습니다")
    if txn.status in (TransactionStatus.CANCELLED, TransactionStatus.HIDDEN):
        raise HTTPException(status_code=400, detail="취소/숨김 이벤트에는 배분할 수 없습니다")
    if txn.status == TransactionStatus.ON_HOLD:
        raise HTTPException(status_code=400, detail="보류 상태에서는 배분할 수 없습니다. 먼저 보류를 해제하세요")
    if txn.status == TransactionStatus.ALLOCATED:
        raise HTTPException(status_code=400, detail="이미 전액 배분된 이벤트입니다")

    remaining = txn.amount - txn.allocated_amount

    # DEPOSIT → SALES 전표, WITHDRAWAL → PURCHASE 전표
    if txn.transaction_type == TransactionType.DEPOSIT:
        target_type = VoucherType.SALES
    else:
        target_type = VoucherType.PURCHASE

    query = (
        select(Voucher)
        .where(
            Voucher.counterparty_id == txn.counterparty_id,
            Voucher.voucher_type == target_type,
            Voucher.settlement_status != SettlementStatus.LOCKED,
            Voucher.payment_status != PaymentStatus.LOCKED,
        )
        .order_by(Voucher.trade_date.asc(), Voucher.created_at.asc())
    )
    if data.voucher_ids:
        query = query.where(Voucher.id.in_(data.voucher_ids))

    result = await db.execute(query)
    vouchers = result.scalars().all()

    allocations = []
    order = (await db.execute(
        select(func.coalesce(func.max(TransactionAllocation.allocation_order), 0))
        .where(TransactionAllocation.transaction_id == txn.id)
    )).scalar() or 0

    # 배분액/레거시/기존 배분 배치 조회 (N+1 → 4쿼리)
    v_ids = [v.id for v in vouchers]
    alloc_map = {}
    receipt_map = {}
    payment_map = {}
    existing_set = set()
    if v_ids:
        alloc_result = await db.execute(
            select(TransactionAllocation.voucher_id, func.coalesce(func.sum(TransactionAllocation.allocated_amount), 0))
            .where(TransactionAllocation.voucher_id.in_(v_ids))
            .group_by(TransactionAllocation.voucher_id)
        )
        alloc_map = {row[0]: row[1] for row in alloc_result.all()}

        receipt_result = await db.execute(
            select(Receipt.voucher_id, func.coalesce(func.sum(Receipt.amount), 0))
            .where(Receipt.voucher_id.in_(v_ids))
            .group_by(Receipt.voucher_id)
        )
        receipt_map = {row[0]: row[1] for row in receipt_result.all()}

        payment_result = await db.execute(
            select(Payment.voucher_id, func.coalesce(func.sum(Payment.amount), 0))
            .where(Payment.voucher_id.in_(v_ids))
            .group_by(Payment.voucher_id)
        )
        payment_map = {row[0]: row[1] for row in payment_result.all()}

        # 이미 이 transaction에 배분된 전표 ID 목록
        existing_result = await db.execute(
            select(TransactionAllocation.voucher_id)
            .where(
                TransactionAllocation.transaction_id == txn.id,
                TransactionAllocation.voucher_id.in_(v_ids),
            )
        )
        existing_set = {row[0] for row in existing_result.all()}

    for v in vouchers:
        if remaining <= 0:
            break

        if v.id in existing_set:
            continue

        already_allocated = alloc_map.get(v.id, Decimal("0"))
        if v.voucher_type == VoucherType.SALES or (hasattr(v.voucher_type, 'value') and v.voucher_type.value == VoucherType.SALES.value):
            legacy = receipt_map.get(v.id, Decimal("0"))
        else:
            legacy = payment_map.get(v.id, Decimal("0"))
        voucher_balance = v.total_amount - already_allocated - legacy

        if voucher_balance <= 0:
            continue

        alloc_amount = min(remaining, voucher_balance)
        order += 1
        allocation = TransactionAllocation(
            transaction_id=txn.id,
            voucher_id=v.id,
            allocated_amount=alloc_amount,
            allocation_order=order,
            created_by=current_user.id,
        )
        db.add(allocation)
        allocations.append((allocation, v))
        remaining -= alloc_amount

    await db.flush()

    # Transaction 상태 업데이트
    await _update_transaction_status(txn, db)

    # 전표 상태 업데이트
    for alloc, v in allocations:
        await _update_voucher_status(v.id, db)

    db.add(AuditLog(
        user_id=current_user.id,
        action=AuditAction.ALLOCATION_AUTO,
        target_type="counterparty_transaction",
        target_id=txn.id,
        after_data={
            "strategy": data.strategy,
            "allocated_count": len(allocations),
            "allocated_total": str(txn.allocated_amount),
        },
    ))

    return [
        AllocationResponse(
            id=a.id,
            transaction_id=a.transaction_id,
            voucher_id=a.voucher_id,
            voucher_number=v.voucher_number,
            voucher_trade_date=v.trade_date,
            voucher_total_amount=v.total_amount,
            allocated_amount=a.allocated_amount,
            allocation_order=a.allocation_order,
            memo=a.memo,
            created_at=a.created_at,
        )
        for a, v in allocations
    ]


@router.post("/{transaction_id}/allocate", response_model=list[AllocationResponse])
async def manual_allocate(
    transaction_id: UUID,
    data: AllocationRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """수동 배분 (전표 지정)"""
    txn = await db.get(CounterpartyTransaction, transaction_id)
    if not txn:
        raise HTTPException(status_code=404, detail="입출금 이벤트를 찾을 수 없습니다")
    if txn.status in (TransactionStatus.CANCELLED, TransactionStatus.HIDDEN, TransactionStatus.ALLOCATED):
        raise HTTPException(status_code=400, detail="배분할 수 없는 상태입니다")
    if txn.status == TransactionStatus.ON_HOLD:
        raise HTTPException(status_code=400, detail="보류 상태에서는 배분할 수 없습니다. 먼저 보류를 해제하세요")

    remaining = txn.amount - txn.allocated_amount
    request_total = sum(item.amount for item in data.allocations)
    if request_total > remaining:
        raise HTTPException(
            status_code=400,
            detail=f"배분 요청액({request_total})이 미배분 잔액({remaining})을 초과합니다"
        )

    # 대상 전표 방향 검증
    expected_type = VoucherType.SALES if txn.transaction_type == TransactionType.DEPOSIT else VoucherType.PURCHASE

    max_order = (await db.execute(
        select(func.coalesce(func.max(TransactionAllocation.allocation_order), 0))
        .where(TransactionAllocation.transaction_id == txn.id)
    )).scalar() or 0

    # 전표 배치 로드 + 배분액 배치 계산
    req_voucher_ids = [item.voucher_id for item in data.allocations]
    v_result = await db.execute(select(Voucher).where(Voucher.id.in_(req_voucher_ids)))
    voucher_map = {v.id: v for v in v_result.scalars().all()}

    alloc_result = await db.execute(
        select(TransactionAllocation.voucher_id, func.coalesce(func.sum(TransactionAllocation.allocated_amount), 0))
        .where(TransactionAllocation.voucher_id.in_(req_voucher_ids))
        .group_by(TransactionAllocation.voucher_id)
    )
    alloc_sum_map = {row[0]: row[1] for row in alloc_result.all()}

    receipt_result = await db.execute(
        select(Receipt.voucher_id, func.coalesce(func.sum(Receipt.amount), 0))
        .where(Receipt.voucher_id.in_(req_voucher_ids))
        .group_by(Receipt.voucher_id)
    )
    receipt_map = {row[0]: row[1] for row in receipt_result.all()}

    payment_result = await db.execute(
        select(Payment.voucher_id, func.coalesce(func.sum(Payment.amount), 0))
        .where(Payment.voucher_id.in_(req_voucher_ids))
        .group_by(Payment.voucher_id)
    )
    payment_map = {row[0]: row[1] for row in payment_result.all()}

    # 이미 배분된 전표 ID
    existing_result = await db.execute(
        select(TransactionAllocation.voucher_id)
        .where(
            TransactionAllocation.transaction_id == txn.id,
            TransactionAllocation.voucher_id.in_(req_voucher_ids),
        )
    )
    existing_set = {row[0] for row in existing_result.all()}

    allocations = []
    for item in data.allocations:
        v = voucher_map.get(item.voucher_id)
        if not v:
            raise HTTPException(status_code=404, detail=f"전표 {item.voucher_id}를 찾을 수 없습니다")
        if v.counterparty_id != txn.counterparty_id:
            raise HTTPException(status_code=400, detail=f"전표 {v.voucher_number}는 다른 거래처 소속입니다")
        if v.voucher_type != expected_type:
            raise HTTPException(
                status_code=400,
                detail=f"전표 타입 불일치: {txn.transaction_type.value}는 {expected_type.value} 전표에만 배분 가능"
            )
        if v.settlement_status == SettlementStatus.LOCKED or v.payment_status == PaymentStatus.LOCKED:
            raise HTTPException(status_code=400, detail=f"마감된 전표 {v.voucher_number}에는 배분할 수 없습니다")

        # 전표 잔액 검증
        already = alloc_sum_map.get(v.id, Decimal("0"))
        if v.voucher_type == VoucherType.SALES or (hasattr(v.voucher_type, 'value') and v.voucher_type.value == VoucherType.SALES.value):
            legacy = receipt_map.get(v.id, Decimal("0"))
        else:
            legacy = payment_map.get(v.id, Decimal("0"))
        voucher_balance = v.total_amount - already - legacy
        if item.amount > voucher_balance:
            raise HTTPException(
                status_code=400,
                detail=f"전표 {v.voucher_number}의 잔액({voucher_balance})보다 배분액({item.amount})이 큽니다"
            )

        # 중복 검증
        if v.id in existing_set:
            raise HTTPException(status_code=400, detail=f"전표 {v.voucher_number}에 이미 배분되어 있습니다")

        max_order += 1
        alloc = TransactionAllocation(
            transaction_id=txn.id,
            voucher_id=v.id,
            allocated_amount=item.amount,
            allocation_order=max_order,
            created_by=current_user.id,
        )
        db.add(alloc)
        allocations.append((alloc, v))

    await db.flush()

    await _update_transaction_status(txn, db)
    for alloc, v in allocations:
        await _update_voucher_status(v.id, db)

    db.add(AuditLog(
        user_id=current_user.id,
        action=AuditAction.ALLOCATION_CREATE,
        target_type="counterparty_transaction",
        target_id=txn.id,
        after_data={
            "allocations": [
                {"voucher_id": str(a.voucher_id), "amount": str(a.allocated_amount)}
                for a, _ in allocations
            ]
        },
    ))

    return [
        AllocationResponse(
            id=a.id,
            transaction_id=a.transaction_id,
            voucher_id=a.voucher_id,
            voucher_number=v.voucher_number,
            voucher_trade_date=v.trade_date,
            voucher_total_amount=v.total_amount,
            allocated_amount=a.allocated_amount,
            allocation_order=a.allocation_order,
            memo=a.memo,
            created_at=a.created_at,
        )
        for a, v in allocations
    ]


@router.delete("/{transaction_id}/allocations/{allocation_id}", status_code=204)
async def delete_allocation(
    transaction_id: UUID,
    allocation_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """배분 삭제"""
    alloc = await db.execute(
        select(TransactionAllocation)
        .where(
            TransactionAllocation.id == allocation_id,
            TransactionAllocation.transaction_id == transaction_id,
        )
    )
    alloc = alloc.scalar_one_or_none()
    if not alloc:
        raise HTTPException(status_code=404, detail="배분 내역을 찾을 수 없습니다")

    voucher_id = alloc.voucher_id
    amount = alloc.allocated_amount

    db.add(AuditLog(
        user_id=current_user.id,
        action=AuditAction.ALLOCATION_DELETE,
        target_type="transaction_allocation",
        target_id=allocation_id,
        before_data={"voucher_id": str(voucher_id), "amount": str(amount)},
    ))

    await db.delete(alloc)
    await db.flush()

    # 상태 재계산
    txn = await db.get(CounterpartyTransaction, transaction_id)
    if txn:
        await _update_transaction_status(txn, db)
    await _update_voucher_status(voucher_id, db)


# =============================================================================
# 거래처 타임라인 / 잔액
# =============================================================================

@router.get("/counterparty/{counterparty_id}/timeline")
async def get_counterparty_timeline(
    counterparty_id: UUID,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """거래처 입출금 타임라인"""
    cp = await db.get(Counterparty, counterparty_id)
    if not cp:
        raise HTTPException(status_code=404, detail="거래처를 찾을 수 없습니다")

    base_filters = [
        CounterpartyTransaction.counterparty_id == counterparty_id,
        CounterpartyTransaction.status.notin_([
            TransactionStatus.CANCELLED,
            TransactionStatus.HIDDEN,
        ]),
    ]
    if date_from:
        base_filters.append(CounterpartyTransaction.transaction_date >= date_from)
    if date_to:
        base_filters.append(CounterpartyTransaction.transaction_date <= date_to)

    # 전체 건수
    count_q = select(func.count(CounterpartyTransaction.id)).where(*base_filters)
    total = (await db.execute(count_q)).scalar() or 0

    query = (
        select(CounterpartyTransaction)
        .where(*base_filters)
        .order_by(
            CounterpartyTransaction.transaction_date.desc(),
            CounterpartyTransaction.created_at.desc(),
        )
        .offset((page - 1) * page_size)
        .limit(page_size)
    )

    result = await db.execute(query)
    txns = result.scalars().all()

    # 배분 건수 일괄 조회
    txn_ids = [t.id for t in txns]
    alloc_counts = {}
    if txn_ids:
        count_result = await db.execute(
            select(
                TransactionAllocation.transaction_id,
                func.count(TransactionAllocation.id),
            )
            .where(TransactionAllocation.transaction_id.in_(txn_ids))
            .group_by(TransactionAllocation.transaction_id)
        )
        alloc_counts = {row[0]: row[1] for row in count_result.all()}

    return {
        "timeline": [
            CounterpartyTimelineItem(
                id=t.id,
                transaction_type=t.transaction_type.value if hasattr(t.transaction_type, 'value') else t.transaction_type,
                transaction_date=t.transaction_date,
                amount=t.amount,
                allocated_amount=t.allocated_amount,
                unallocated_amount=t.amount - t.allocated_amount,
                source=t.source.value if hasattr(t.source, 'value') else t.source,
                status=t.status.value if hasattr(t.status, 'value') else t.status,
                memo=t.memo,
                allocation_count=alloc_counts.get(t.id, 0),
                created_at=t.created_at,
            )
            for t in txns
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.get("/counterparty/{counterparty_id}/balance", response_model=CounterpartyBalanceSummary)
async def get_counterparty_balance(
    counterparty_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """거래처 잔액 요약"""
    cp = await db.get(Counterparty, counterparty_id)
    if not cp:
        raise HTTPException(status_code=404, detail="거래처를 찾을 수 없습니다")

    # 입금 합계
    deposit_result = await db.execute(
        select(
            func.coalesce(func.sum(CounterpartyTransaction.amount), 0),
            func.coalesce(func.sum(CounterpartyTransaction.allocated_amount), 0),
        )
        .where(
            CounterpartyTransaction.counterparty_id == counterparty_id,
            CounterpartyTransaction.transaction_type == TransactionType.DEPOSIT,
            CounterpartyTransaction.status.notin_([TransactionStatus.CANCELLED, TransactionStatus.HIDDEN]),
        )
    )
    dep_row = deposit_result.one()
    total_deposits = dep_row[0]
    total_alloc_deposits = dep_row[1]

    # 출금 합계
    withdrawal_result = await db.execute(
        select(
            func.coalesce(func.sum(CounterpartyTransaction.amount), 0),
            func.coalesce(func.sum(CounterpartyTransaction.allocated_amount), 0),
        )
        .where(
            CounterpartyTransaction.counterparty_id == counterparty_id,
            CounterpartyTransaction.transaction_type == TransactionType.WITHDRAWAL,
            CounterpartyTransaction.status.notin_([TransactionStatus.CANCELLED, TransactionStatus.HIDDEN]),
        )
    )
    wd_row = withdrawal_result.one()
    total_withdrawals = wd_row[0]
    total_alloc_withdrawals = wd_row[1]

    # 미수/미지급 (전표 기준)
    receivable = (await db.execute(
        select(func.coalesce(func.sum(Voucher.total_amount), 0))
        .where(
            Voucher.counterparty_id == counterparty_id,
            Voucher.voucher_type == VoucherType.SALES,
            Voucher.settlement_status.in_([SettlementStatus.OPEN, SettlementStatus.SETTLING]),
        )
    )).scalar() or Decimal("0")

    payable = (await db.execute(
        select(func.coalesce(func.sum(Voucher.total_amount), 0))
        .where(
            Voucher.counterparty_id == counterparty_id,
            Voucher.voucher_type == VoucherType.PURCHASE,
            Voucher.payment_status.in_([PaymentStatus.UNPAID, PaymentStatus.PARTIAL]),
        )
    )).scalar() or Decimal("0")

    return CounterpartyBalanceSummary(
        counterparty_id=counterparty_id,
        counterparty_name=cp.name,
        total_deposits=total_deposits,
        total_withdrawals=total_withdrawals,
        total_allocated_deposits=total_alloc_deposits,
        total_allocated_withdrawals=total_alloc_withdrawals,
        unallocated_deposits=total_deposits - total_alloc_deposits,
        unallocated_withdrawals=total_withdrawals - total_alloc_withdrawals,
        total_receivable=receivable,
        total_payable=payable,
    )


# =============================================================================
# 엑셀 업로드 (다량 등록)
# =============================================================================

def _parse_korean_date(raw: str, year: int) -> Optional[date]:
    """'10월 17일' 또는 '10/17' 형식을 date로 변환"""
    if not raw:
        return None
    raw = str(raw).strip()
    # 10월 17일
    m = re.match(r"(\d{1,2})월\s*(\d{1,2})일?", raw)
    if m:
        return date(year, int(m.group(1)), int(m.group(2)))
    # 10/17 or 10-17
    m = re.match(r"(\d{1,2})[/\-](\d{1,2})", raw)
    if m:
        return date(year, int(m.group(1)), int(m.group(2)))
    # 2025-10-17 full date
    m = re.match(r"(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})", raw)
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


def _parse_amount(raw) -> Optional[Decimal]:
    """금액 파싱: 쉼표/공백 제거, 숫자 변환"""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        if raw == 0:
            return None
        return Decimal(str(raw))
    s = str(raw).strip().replace(",", "").replace(" ", "").replace("-", "")
    if not s:
        return None
    try:
        val = Decimal(s)
        return val if val > 0 else None
    except InvalidOperation:
        return None


REQUIRED_COLUMNS = ["날짜", "구분", "거래처", "거래내역", "차변(입금)", "대변(출금)"]
VALID_CATEGORIES = {"판매", "매입"}
SKIPPED_CATEGORIES = {"대체", "비용", "기타", "이체"}


def _find_header_row(ws) -> Optional[int]:
    """날짜/구분/거래처 컬럼이 있는 헤더 행 탐색"""
    for row_idx in range(1, min(ws.max_row + 1, 20)):
        cells = [str(ws.cell(row=row_idx, column=c).value or "").strip() for c in range(1, min(ws.max_column + 1, 15))]
        joined = " ".join(cells).lower()
        if "날짜" in joined and ("거래처" in joined or "업체" in joined):
            return row_idx
    return None


def _detect_columns(ws, header_row: int) -> dict:
    """헤더 행에서 컬럼 인덱스 매핑 (새 양식: 날짜/구분/거래처/거래내역/차변(입금)/대변(출금)/잔액)"""
    cols = {}
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=c).value or "").strip()
        lower = val.lower().replace(" ", "")
        if "날짜" in lower and "date" not in cols:
            cols["date"] = c
        elif "구분" in lower and "category" not in cols:
            cols["category"] = c
        elif ("거래처" in lower or "업체" in lower) and "counterparty" not in cols:
            cols["counterparty"] = c
        elif ("거래내역" in lower or "내역" in lower) and "description" not in cols:
            cols["description"] = c
        elif ("차변" in lower or "입금" in lower) and "deposit" not in cols:
            cols["deposit"] = c
        elif ("대변" in lower or "출금" in lower) and "withdrawal" not in cols:
            cols["withdrawal"] = c
        elif "잔액" in lower and "balance" not in cols:
            cols["balance"] = c
    return cols


def _is_summary_row(cells: dict) -> bool:
    """합계/소계 행 감지"""
    cp = str(cells.get("counterparty", "") or "").strip()
    if not cp:
        return True
    keywords = ["합계", "소계", "total", "sum", "계"]
    return any(k in cp.lower() for k in keywords)


async def _match_counterparty(name: str, db: AsyncSession) -> Optional[UUID]:
    """거래처 이름 → ID 매칭 (alias 우선, 표준명 fallback)"""
    # 1) alias
    result = await db.execute(
        select(CounterpartyAlias.counterparty_id).where(
            CounterpartyAlias.alias_name == name
        ).limit(1)
    )
    cp_id = result.scalar_one_or_none()
    if cp_id:
        return cp_id
    # 2) standard name
    result = await db.execute(
        select(Counterparty.id).where(
            Counterparty.name == name,
            Counterparty.is_active == True,
        ).limit(1)
    )
    return result.scalar_one_or_none()


@router.post("/upload/preview")
async def preview_transaction_upload(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """엑셀 파일을 파싱하여 미리보기 반환"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.")

    import openpyxl
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="엑셀 파일을 읽을 수 없습니다.")

    ws = wb.active
    header_row = _find_header_row(ws)
    if not header_row:
        raise HTTPException(status_code=400, detail="헤더 행을 찾을 수 없습니다. '날짜', '거래처' 컬럼이 필요합니다.")

    cols = _detect_columns(ws, header_row)

    # ── 양식 검증: 필수 컬럼 체크 ──
    missing = []
    if "date" not in cols:
        missing.append("날짜")
    if "category" not in cols:
        missing.append("구분")
    if "counterparty" not in cols:
        missing.append("거래처")
    if "deposit" not in cols:
        missing.append("차변(입금)")
    if "withdrawal" not in cols:
        missing.append("대변(출금)")

    if missing:
        raise HTTPException(
            status_code=400,
            detail=(
                f"엑셀 양식이 올바르지 않습니다. "
                f"누락된 필수 컬럼: {', '.join(missing)}. "
                f"필수 컬럼: {', '.join(REQUIRED_COLUMNS)}. "
                f"구분 컬럼에는 '판매' 또는 '매입' 값이 필요합니다."
            ),
        )

    current_year = datetime.now().year
    rows = []
    unmatched_set = set()
    total_deposit = Decimal("0")
    total_withdrawal = Decimal("0")
    valid_count = 0
    error_count = 0
    unmatched_count = 0
    skipped_count = 0
    last_date_val = None

    for row_idx in range(header_row + 1, ws.max_row + 1):
        date_raw = ws.cell(row=row_idx, column=cols["date"]).value
        category_raw = ws.cell(row=row_idx, column=cols["category"]).value
        cp_raw = ws.cell(row=row_idx, column=cols["counterparty"]).value
        desc_raw = ws.cell(row=row_idx, column=cols["description"]).value if cols.get("description") else None
        dep_raw = ws.cell(row=row_idx, column=cols["deposit"]).value if cols.get("deposit") else None
        wth_raw = ws.cell(row=row_idx, column=cols["withdrawal"]).value if cols.get("withdrawal") else None
        bal_raw = ws.cell(row=row_idx, column=cols["balance"]).value if cols.get("balance") else None

        cell_data = {"counterparty": cp_raw}
        if _is_summary_row(cell_data):
            continue

        category = str(category_raw).strip() if category_raw else ""

        # 구분 값이 비어 있으면 행 건너뜀
        if not category:
            continue

        # 날짜 파싱 (빈 날짜 → 이전 행 날짜 유지)
        if date_raw is not None:
            if isinstance(date_raw, datetime):
                parsed_date = date_raw.date()
            elif isinstance(date_raw, date):
                parsed_date = date_raw
            else:
                parsed_date = _parse_korean_date(str(date_raw), current_year)
            if parsed_date:
                last_date_val = parsed_date
        parsed_date = last_date_val

        cp_name = str(cp_raw).strip() if cp_raw else ""
        description = str(desc_raw).strip() if desc_raw else ""
        deposit_amount = _parse_amount(dep_raw)
        withdrawal_amount = _parse_amount(wth_raw)
        balance = _parse_amount(bal_raw)

        # ── 구분 기반 처리 ──
        # 대체/비용/기타 등 판매·매입 외 구분은 건너뜀(skipped)
        if category not in VALID_CATEGORIES:
            skipped_count += 1
            rows.append({
                "row_number": row_idx - header_row,
                "transaction_date": parsed_date.isoformat() if parsed_date else None,
                "counterparty_name": cp_name,
                "counterparty_id": None,
                "category": category,
                "description": description,
                "transaction_type": None,
                "amount": float(deposit_amount or withdrawal_amount or 0),
                "balance": float(balance) if balance else None,
                "status": "skipped",
                "message": f"'{category}' 구분은 처리 대상이 아닙니다 (판매/매입만 처리)",
            })
            continue

        # 판매: 차변(입금) 사용 → deposit
        if category == "판매":
            if not deposit_amount:
                error_count += 1
                rows.append({
                    "row_number": row_idx - header_row,
                    "transaction_date": parsed_date.isoformat() if parsed_date else None,
                    "counterparty_name": cp_name,
                    "counterparty_id": None,
                    "category": category,
                    "description": description,
                    "transaction_type": "deposit",
                    "amount": 0,
                    "balance": float(balance) if balance else None,
                    "status": "error",
                    "message": "판매 구분이지만 차변(입금) 금액이 없습니다",
                })
                continue
            txn_type = "deposit"
            amount = deposit_amount

        # 매입: 대변(출금) 사용 → withdrawal
        else:
            if not withdrawal_amount:
                error_count += 1
                rows.append({
                    "row_number": row_idx - header_row,
                    "transaction_date": parsed_date.isoformat() if parsed_date else None,
                    "counterparty_name": cp_name,
                    "counterparty_id": None,
                    "category": category,
                    "description": description,
                    "transaction_type": "withdrawal",
                    "amount": 0,
                    "balance": float(balance) if balance else None,
                    "status": "error",
                    "message": "매입 구분이지만 대변(출금) 금액이 없습니다",
                })
                continue
            txn_type = "withdrawal"
            amount = withdrawal_amount

        # ── 검증 ──
        row_status = "ok"
        message = None
        cp_id = None

        if not parsed_date:
            row_status = "error"
            message = "날짜를 파싱할 수 없습니다"
            error_count += 1
        elif not cp_name:
            row_status = "error"
            message = "거래처명이 비어있습니다"
            error_count += 1
        else:
            cp_id = await _match_counterparty(cp_name, db)
            if not cp_id:
                row_status = "unmatched"
                message = "거래처를 찾을 수 없습니다"
                unmatched_set.add(cp_name)
                unmatched_count += 1
            else:
                valid_count += 1
                if txn_type == "deposit":
                    total_deposit += amount
                else:
                    total_withdrawal += amount

        rows.append({
            "row_number": row_idx - header_row,
            "transaction_date": parsed_date.isoformat() if parsed_date else None,
            "counterparty_name": cp_name,
            "counterparty_id": str(cp_id) if cp_id else None,
            "category": category,
            "description": description,
            "transaction_type": txn_type,
            "amount": float(amount),
            "balance": float(balance) if balance else None,
            "status": row_status,
            "message": message,
        })

    wb.close()

    return {
        "rows": rows,
        "summary": {
            "total": len(rows),
            "valid": valid_count,
            "error": error_count,
            "unmatched": unmatched_count,
            "skipped": skipped_count,
            "total_deposit": float(total_deposit),
            "total_withdrawal": float(total_withdrawal),
        },
        "unmatched_counterparties": sorted(unmatched_set),
    }


@router.post("/upload/confirm")
async def confirm_transaction_upload(
    payload: dict = Body(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """미리보기에서 확인된 입출금 건을 일괄 등록"""
    transactions = payload.get("transactions", [])
    if not transactions:
        raise HTTPException(status_code=400, detail="등록할 거래가 없습니다.")

    created_count = 0
    skipped_count = 0
    errors = []

    for idx, txn in enumerate(transactions):
        try:
            cp_id = UUID(txn["counterparty_id"])
            txn_date = date.fromisoformat(txn["transaction_date"])
            txn_type = txn["transaction_type"]
            amount = Decimal(str(txn["amount"]))
            memo = txn.get("memo", "")

            # 중복 체크
            dup_check = await db.execute(
                select(CounterpartyTransaction.id).where(
                    CounterpartyTransaction.counterparty_id == cp_id,
                    CounterpartyTransaction.transaction_date == txn_date,
                    CounterpartyTransaction.transaction_type == (
                        TransactionType.DEPOSIT if txn_type == "deposit" else TransactionType.WITHDRAWAL
                    ),
                    CounterpartyTransaction.amount == amount,
                    CounterpartyTransaction.status != TransactionStatus.CANCELLED,
                ).limit(1)
            )
            if dup_check.scalar_one_or_none():
                skipped_count += 1
                continue

            new_txn = CounterpartyTransaction(
                counterparty_id=cp_id,
                transaction_type=TransactionType.DEPOSIT if txn_type == "deposit" else TransactionType.WITHDRAWAL,
                transaction_date=txn_date,
                amount=amount,
                allocated_amount=Decimal("0"),
                memo=memo or None,
                source=TransactionSource.MANUAL,
                status=TransactionStatus.PENDING,
                created_by=current_user.id,
            )
            db.add(new_txn)
            created_count += 1

        except Exception as e:
            errors.append({"row": idx + 1, "error": str(e)})

    await db.commit()

    return {
        "created_count": created_count,
        "skipped_count": skipped_count,
        "errors": errors,
    }
