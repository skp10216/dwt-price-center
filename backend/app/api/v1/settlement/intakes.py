"""
정산 도메인 - 반입 내역 CRUD API
margin/margin_rate는 서버에서 계산하여 응답 (DB에 저장하지 않음)
current_status/intake_type은 enum 기반
"""

import uuid
import io
from datetime import date, datetime
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Body
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, func, or_, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.api.deps import get_settlement_user
from app.models.user import User
from app.models.intake_item import IntakeItem
from app.models.counterparty import Counterparty
from app.models.enums import AuditAction, IntakeStatus, IntakeType
from app.models.audit_log import AuditLog

router = APIRouter()


# ─── Pydantic ─────────────────────────────────────────────────────

class IntakeItemUpdate(BaseModel):
    pg_no: Optional[str] = None
    model_name: Optional[str] = None
    serial_number: Optional[str] = None
    actual_purchase_price: Optional[float] = None
    intake_price: Optional[float] = None
    intake_type: Optional[str] = None
    current_status: Optional[str] = None
    remarks: Optional[str] = None
    memo: Optional[str] = None


class StatusChangeRequest(BaseModel):
    status: str


# ─── 헬퍼 ─────────────────────────────────────────────────────────

def _apply_filters(query, date_from, date_to, counterparty_id, status, intake_type, search):
    if date_from:
        query = query.where(IntakeItem.intake_date >= date_from)
    if date_to:
        query = query.where(IntakeItem.intake_date <= date_to)
    if counterparty_id:
        query = query.where(IntakeItem.counterparty_id == counterparty_id)
    if status:
        try:
            query = query.where(IntakeItem.current_status == IntakeStatus(status))
        except ValueError:
            pass
    if intake_type:
        try:
            query = query.where(IntakeItem.intake_type == IntakeType(intake_type))
        except ValueError:
            pass
    if search:
        term = f"%{search}%"
        query = query.where(or_(
            IntakeItem.serial_number.ilike(term),
            IntakeItem.pg_no.ilike(term),
            IntakeItem.model_name.ilike(term),
            IntakeItem.slip_number.ilike(term),
            IntakeItem.remarks.ilike(term),
        ))
    return query


def _to_response(item: IntakeItem) -> dict:
    app_val = float(item.actual_purchase_price or 0)
    ip_val = float(item.intake_price or 0)
    margin = app_val - ip_val
    margin_rate = (margin / app_val * 100) if app_val != 0 else 0

    return {
        "id": item.id,
        "intake_date": item.intake_date.isoformat(),
        "slip_number": item.slip_number,
        "counterparty_id": item.counterparty_id,
        "counterparty_name": item.counterparty.name if item.counterparty else None,
        "pg_no": item.pg_no,
        "model_name": item.model_name,
        "serial_number": item.serial_number,
        "purchase_date": item.purchase_date.isoformat() if item.purchase_date else None,
        "purchase_counterparty_id": item.purchase_counterparty_id,
        "purchase_counterparty_name": item.purchase_counterparty.name if item.purchase_counterparty else None,
        "actual_purchase_price": app_val,
        "intake_price": ip_val,
        "margin": round(margin, 2),
        "margin_rate": round(margin_rate, 2),
        "intake_type": item.intake_type.value if item.intake_type else "normal",
        "current_status": item.current_status.value if item.current_status else "received",
        "remarks": item.remarks,
        "memo": item.memo,
        "is_locked": item.is_locked,
        "source_voucher_id": item.source_voucher_id,
        "created_at": item.created_at.isoformat() if item.created_at else None,
        "updated_at": item.updated_at.isoformat() if item.updated_at else None,
    }


# ─── 목록 ──────────────────────────────────────────────────────────

@router.get("", response_model=dict)
async def list_intake_items(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    counterparty_id: Optional[uuid.UUID] = Query(None),
    status: Optional[str] = Query(None, alias="current_status"),
    intake_type: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    sort_by: str = Query("intake_date"),
    sort_order: str = Query("desc"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_settlement_user),
):
    """반입 내역 목록"""
    base = select(IntakeItem).options(
        selectinload(IntakeItem.counterparty),
        selectinload(IntakeItem.purchase_counterparty),
    )
    base = _apply_filters(base, date_from, date_to, counterparty_id, status, intake_type, search)

    sort_col = getattr(IntakeItem, sort_by, IntakeItem.intake_date)
    base = base.order_by(sort_col.asc() if sort_order == "asc" else sort_col.desc())

    count_q = select(func.count(IntakeItem.id))
    count_q = _apply_filters(count_q, date_from, date_to, counterparty_id, status, intake_type, search)
    total = (await db.execute(count_q)).scalar() or 0

    base = base.offset((page - 1) * page_size).limit(page_size)
    items = (await db.execute(base)).scalars().all()

    return {"items": [_to_response(i) for i in items], "total": total, "page": page, "page_size": page_size}


# ─── 합계 (margin 서버 계산) ────────────────────────────────────────

@router.get("/summary", response_model=dict)
async def get_intake_summary(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    counterparty_id: Optional[uuid.UUID] = Query(None),
    status: Optional[str] = Query(None, alias="current_status"),
    intake_type: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_settlement_user),
):
    """반입 내역 합계 — margin은 DB expression으로 계산"""
    query = select(
        func.count(IntakeItem.id).label("total_count"),
        func.coalesce(func.sum(IntakeItem.actual_purchase_price), 0).label("total_actual"),
        func.coalesce(func.sum(IntakeItem.intake_price), 0).label("total_intake"),
    )
    query = _apply_filters(query, date_from, date_to, counterparty_id, status, intake_type, search)
    row = (await db.execute(query)).one()

    total_actual = float(row.total_actual)
    total_intake = float(row.total_intake)
    total_margin = total_actual - total_intake
    avg_margin_rate = (total_margin / total_actual * 100) if total_actual != 0 else 0

    return {
        "total_count": row.total_count,
        "total_actual_purchase_price": round(total_actual, 2),
        "total_intake_price": round(total_intake, 2),
        "total_margin": round(total_margin, 2),
        "avg_margin_rate": round(avg_margin_rate, 2),
    }


# ─── 상세 ──────────────────────────────────────────────────────────

@router.get("/{item_id}", response_model=dict)
async def get_intake_item(
    item_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_settlement_user),
):
    result = await db.execute(
        select(IntakeItem).options(
            selectinload(IntakeItem.counterparty),
            selectinload(IntakeItem.purchase_counterparty),
        ).where(IntakeItem.id == item_id)
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="반입 내역을 찾을 수 없습니다")
    return _to_response(item)


# ─── 수정 ──────────────────────────────────────────────────────────

@router.patch("/{item_id}", response_model=dict)
async def update_intake_item(
    item_id: uuid.UUID,
    body: IntakeItemUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_settlement_user),
):
    item = await db.get(IntakeItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="반입 내역을 찾을 수 없습니다")
    if item.is_locked:
        raise HTTPException(status_code=400, detail="마감된 반입 내역은 수정할 수 없습니다")

    update_data = body.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if field in ("actual_purchase_price", "intake_price"):
            setattr(item, field, Decimal(str(value)) if value is not None else Decimal("0"))
        elif field == "current_status" and value:
            try:
                item.current_status = IntakeStatus(value)
            except ValueError:
                raise HTTPException(status_code=400, detail=f"잘못된 상태값: {value}")
        elif field == "intake_type" and value:
            try:
                item.intake_type = IntakeType(value)
            except ValueError:
                raise HTTPException(status_code=400, detail=f"잘못된 반입구분: {value}")
        else:
            setattr(item, field, value)

    db.add(AuditLog(
        user_id=current_user.id, action=AuditAction.INTAKE_ITEM_UPDATE,
        target_type="intake_item", target_id=item.id, after_data=update_data,
    ))
    await db.flush()
    return {"message": "수정 완료", "id": str(item_id)}


# ─── 상태 변경 ────────────────────────────────────────────────────

@router.patch("/{item_id}/status", response_model=dict)
async def change_intake_status(
    item_id: uuid.UUID,
    body: StatusChangeRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_settlement_user),
):
    item = await db.get(IntakeItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="반입 내역을 찾을 수 없습니다")
    if item.is_locked:
        raise HTTPException(status_code=400, detail="마감된 반입 내역은 상태 변경할 수 없습니다")

    try:
        new_status = IntakeStatus(body.status)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"잘못된 상태값: {body.status}")

    old_status = item.current_status.value if item.current_status else None
    item.current_status = new_status

    db.add(AuditLog(
        user_id=current_user.id, action=AuditAction.INTAKE_ITEM_STATUS_CHANGE,
        target_type="intake_item", target_id=item.id,
        before_data={"current_status": old_status},
        after_data={"current_status": new_status.value},
    ))
    await db.flush()
    return {"message": "상태 변경 완료", "id": str(item_id), "new_status": new_status.value}


# ─── 삭제 ──────────────────────────────────────────────────────────

@router.delete("/{item_id}", response_model=dict)
async def delete_intake_item(
    item_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_settlement_user),
):
    item = await db.get(IntakeItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="반입 내역을 찾을 수 없습니다")
    if item.is_locked:
        raise HTTPException(status_code=400, detail="마감된 반입 내역은 삭제할 수 없습니다")

    db.add(AuditLog(
        user_id=current_user.id, action=AuditAction.INTAKE_ITEM_DELETE,
        target_type="intake_item", target_id=item.id,
    ))
    await db.delete(item)
    await db.flush()
    return {"message": "삭제 완료", "id": str(item_id)}


# ─── 엑셀 내보내기 ─────────────────────────────────────────────────

@router.get("/export/excel")
async def export_intake_items_excel(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    counterparty_id: Optional[uuid.UUID] = Query(None),
    status: Optional[str] = Query(None, alias="current_status"),
    intake_type: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_settlement_user),
):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    query = select(IntakeItem).options(
        selectinload(IntakeItem.counterparty),
        selectinload(IntakeItem.purchase_counterparty),
    )
    query = _apply_filters(query, date_from, date_to, counterparty_id, status, intake_type, search)
    query = query.order_by(IntakeItem.intake_date.desc())
    items = (await db.execute(query)).scalars().all()

    STATUS_LABEL = {"received": "반입", "in_stock": "재고", "sold": "판매완료", "hold": "보류", "excluded": "제외"}
    TYPE_LABEL = {"normal": "일반", "return_intake": "재반입", "transfer": "이관", "other": "기타"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "반입내역"

    headers = ["반입일", "전표번호", "반입처", "P/G No", "모델명", "일련번호",
               "매입일", "매입처", "실매입가", "반입가", "마진", "반입구분", "현상태",
               "특이사항", "비고"]

    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hfont = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill, c.font, c.alignment, c.border = hfill, hfont, Alignment(horizontal="center", vertical="center"), border

    for ri, item in enumerate(items, 2):
        app_val = float(item.actual_purchase_price or 0)
        ip_val = float(item.intake_price or 0)
        vals = [
            item.intake_date.isoformat() if item.intake_date else "",
            item.slip_number,
            item.counterparty.name if item.counterparty else "",
            item.pg_no or "",
            item.model_name or "",
            item.serial_number or "",
            item.purchase_date.isoformat() if item.purchase_date else "",
            item.purchase_counterparty.name if item.purchase_counterparty else "",
            app_val, ip_val, round(app_val - ip_val, 2),
            TYPE_LABEL.get(item.intake_type.value, item.intake_type.value) if item.intake_type else "",
            STATUS_LABEL.get(item.current_status.value, item.current_status.value) if item.current_status else "",
            item.remarks or "", item.memo or "",
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = border
            if ci in (9, 10, 11):
                c.number_format = "#,##0"

    for col in ws.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=반입내역_{date.today().isoformat()}.xlsx"},
    )
