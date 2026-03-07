"""
정산 도메인 - 반품 내역 CRUD API
목록 조회(필터/검색/페이지네이션) / 상세 / 수정 / 삭제 / 엑셀 내보내기 / 합계
"""

import uuid
import io
from datetime import date, datetime
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import select, func, or_, and_, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.api.deps import get_settlement_user
from app.models.user import User
from app.models.return_item import ReturnItem
from app.models.counterparty import Counterparty
from app.models.enums import AuditAction
from app.models.audit_log import AuditLog

router = APIRouter()


# ─── Pydantic 스키마 ──────────────────────────────────────────────

class ReturnItemResponse(BaseModel):
    id: uuid.UUID
    return_date: date
    slip_number: str
    counterparty_id: uuid.UUID
    counterparty_name: Optional[str] = None
    pg_no: Optional[str] = None
    model_name: Optional[str] = None
    serial_number: Optional[str] = None
    imei: Optional[str] = None
    color: Optional[str] = None
    purchase_cost: float = 0
    purchase_deduction: float = 0
    return_amount: float = 0
    as_cost: float = 0
    remarks: Optional[str] = None
    memo: Optional[str] = None
    is_locked: bool = False
    source_voucher_id: Optional[uuid.UUID] = None
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


class ReturnItemUpdate(BaseModel):
    pg_no: Optional[str] = None
    model_name: Optional[str] = None
    serial_number: Optional[str] = None
    imei: Optional[str] = None
    color: Optional[str] = None
    purchase_cost: Optional[float] = None
    purchase_deduction: Optional[float] = None
    return_amount: Optional[float] = None
    as_cost: Optional[float] = None
    remarks: Optional[str] = None
    memo: Optional[str] = None


class ReturnSummary(BaseModel):
    total_count: int = 0
    total_purchase_cost: float = 0
    total_purchase_deduction: float = 0
    total_return_amount: float = 0
    total_as_cost: float = 0


# ─── 헬퍼 ─────────────────────────────────────────────────────────

def _build_return_query_filters(
    query,
    date_from: Optional[date],
    date_to: Optional[date],
    counterparty_id: Optional[uuid.UUID],
    search: Optional[str],
):
    """공통 필터 적용"""
    if date_from:
        query = query.where(ReturnItem.return_date >= date_from)
    if date_to:
        query = query.where(ReturnItem.return_date <= date_to)
    if counterparty_id:
        query = query.where(ReturnItem.counterparty_id == counterparty_id)
    if search:
        search_term = f"%{search}%"
        query = query.where(
            or_(
                ReturnItem.imei.ilike(search_term),
                ReturnItem.serial_number.ilike(search_term),
                ReturnItem.pg_no.ilike(search_term),
                ReturnItem.model_name.ilike(search_term),
                ReturnItem.slip_number.ilike(search_term),
                ReturnItem.remarks.ilike(search_term),
            )
        )
    return query


def _item_to_response(item: ReturnItem) -> dict:
    data = {
        "id": item.id,
        "return_date": item.return_date.isoformat(),
        "slip_number": item.slip_number,
        "counterparty_id": item.counterparty_id,
        "counterparty_name": item.counterparty.name if item.counterparty else None,
        "pg_no": item.pg_no,
        "model_name": item.model_name,
        "serial_number": item.serial_number,
        "imei": item.imei,
        "color": item.color,
        "purchase_cost": float(item.purchase_cost or 0),
        "purchase_deduction": float(item.purchase_deduction or 0),
        "return_amount": float(item.return_amount or 0),
        "as_cost": float(item.as_cost or 0),
        "remarks": item.remarks,
        "memo": item.memo,
        "is_locked": item.is_locked,
        "source_voucher_id": item.source_voucher_id,
        "created_at": item.created_at.isoformat() if item.created_at else None,
        "updated_at": item.updated_at.isoformat() if item.updated_at else None,
    }
    return data


# ─── 목록 조회 ─────────────────────────────────────────────────────

@router.get("", response_model=dict)
async def list_return_items(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    counterparty_id: Optional[uuid.UUID] = Query(None),
    search: Optional[str] = Query(None, description="IMEI/일련번호/P/G No/모델명 검색"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    sort_by: str = Query("return_date", description="정렬 필드"),
    sort_order: str = Query("desc", description="asc/desc"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_settlement_user),
):
    """반품 내역 목록 조회 (필터/검색/페이지네이션)"""
    base_query = select(ReturnItem).options(
        selectinload(ReturnItem.counterparty)
    )
    base_query = _build_return_query_filters(base_query, date_from, date_to, counterparty_id, search)

    # 정렬
    sort_col = getattr(ReturnItem, sort_by, ReturnItem.return_date)
    if sort_order == "asc":
        base_query = base_query.order_by(sort_col.asc())
    else:
        base_query = base_query.order_by(sort_col.desc())

    # 카운트
    count_base = select(func.count(ReturnItem.id))
    count_base = _build_return_query_filters(count_base, date_from, date_to, counterparty_id, search)
    total = (await db.execute(count_base)).scalar() or 0

    # 페이지네이션
    base_query = base_query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(base_query)
    items = result.scalars().all()

    return {
        "items": [_item_to_response(i) for i in items],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


# ─── 합계 조회 ─────────────────────────────────────────────────────

@router.get("/summary", response_model=ReturnSummary)
async def get_return_summary(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    counterparty_id: Optional[uuid.UUID] = Query(None),
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_settlement_user),
):
    """반품 내역 합계 (필터 적용)"""
    query = select(
        func.count(ReturnItem.id).label("total_count"),
        func.coalesce(func.sum(ReturnItem.purchase_cost), 0).label("total_purchase_cost"),
        func.coalesce(func.sum(ReturnItem.purchase_deduction), 0).label("total_purchase_deduction"),
        func.coalesce(func.sum(ReturnItem.return_amount), 0).label("total_return_amount"),
        func.coalesce(func.sum(ReturnItem.as_cost), 0).label("total_as_cost"),
    )
    query = _build_return_query_filters(query, date_from, date_to, counterparty_id, search)
    result = await db.execute(query)
    row = result.one()

    return ReturnSummary(
        total_count=row.total_count,
        total_purchase_cost=float(row.total_purchase_cost),
        total_purchase_deduction=float(row.total_purchase_deduction),
        total_return_amount=float(row.total_return_amount),
        total_as_cost=float(row.total_as_cost),
    )


# ─── 상세 조회 ─────────────────────────────────────────────────────

@router.get("/{item_id}", response_model=dict)
async def get_return_item(
    item_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_settlement_user),
):
    """반품 내역 상세"""
    result = await db.execute(
        select(ReturnItem)
        .options(selectinload(ReturnItem.counterparty))
        .where(ReturnItem.id == item_id)
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="반품 내역을 찾을 수 없습니다")
    return _item_to_response(item)


# ─── 수정 ──────────────────────────────────────────────────────────

@router.patch("/{item_id}", response_model=dict)
async def update_return_item(
    item_id: uuid.UUID,
    body: ReturnItemUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_settlement_user),
):
    """반품 내역 수정 (잠금 시 불가)"""
    item = await db.get(ReturnItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="반품 내역을 찾을 수 없습니다")
    if item.is_locked:
        raise HTTPException(status_code=400, detail="마감된 반품 내역은 수정할 수 없습니다")

    update_data = body.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if field in ("purchase_cost", "purchase_deduction", "return_amount", "as_cost"):
            setattr(item, field, Decimal(str(value)) if value is not None else Decimal("0"))
        else:
            setattr(item, field, value)

    db.add(AuditLog(
        user_id=current_user.id,
        action=AuditAction.RETURN_ITEM_UPDATE,
        target_type="return_item",
        target_id=item.id,
        after_data=update_data,
    ))
    await db.flush()
    return {"message": "수정 완료", "id": str(item_id)}


# ─── 삭제 ──────────────────────────────────────────────────────────

@router.delete("/{item_id}", response_model=dict)
async def delete_return_item(
    item_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_settlement_user),
):
    """반품 내역 삭제 (잠금 시 불가)"""
    item = await db.get(ReturnItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="반품 내역을 찾을 수 없습니다")
    if item.is_locked:
        raise HTTPException(status_code=400, detail="마감된 반품 내역은 삭제할 수 없습니다")

    db.add(AuditLog(
        user_id=current_user.id,
        action=AuditAction.RETURN_ITEM_DELETE,
        target_type="return_item",
        target_id=item.id,
    ))
    await db.delete(item)
    await db.flush()
    return {"message": "삭제 완료", "id": str(item_id)}


# ─── 엑셀 내보내기 ─────────────────────────────────────────────────

@router.get("/export/excel")
async def export_return_items_excel(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    counterparty_id: Optional[uuid.UUID] = Query(None),
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_settlement_user),
):
    """반품 내역 엑셀 다운로드"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    query = select(ReturnItem).options(selectinload(ReturnItem.counterparty))
    query = _build_return_query_filters(query, date_from, date_to, counterparty_id, search)
    query = query.order_by(ReturnItem.return_date.desc())
    result = await db.execute(query)
    items = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "반품내역"

    headers = ["반품일", "전표번호", "반품처", "P/G No", "모델명", "일련번호",
               "IMEI", "색상", "매입원가", "매입차감", "반품금액", "A/S금액",
               "특이사항", "비고"]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, item in enumerate(items, 2):
        vals = [
            item.return_date.isoformat() if item.return_date else "",
            item.slip_number,
            item.counterparty.name if item.counterparty else "",
            item.pg_no or "",
            item.model_name or "",
            item.serial_number or "",
            item.imei or "",
            item.color or "",
            float(item.purchase_cost or 0),
            float(item.purchase_deduction or 0),
            float(item.return_amount or 0),
            float(item.as_cost or 0),
            item.remarks or "",
            item.memo or "",
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx >= 9 and col_idx <= 12:
                cell.number_format = "#,##0"

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"반품내역_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
